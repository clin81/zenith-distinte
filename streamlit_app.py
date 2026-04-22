import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from streamlit_gsheets import GSheetsConnection

# --- CONFIGURAZIONE PAGINA ---
st.set_page_config(page_title="Zenith Prato - Gestione Distinte", layout="wide")
TEMPLATE_FILE = "distinta_vuota.xlsx"

conn = st.connection("gsheets", type=GSheetsConnection)

# --- FUNZIONE CON CACHE (600 secondi = 10 minuti) ---
@st.cache_data(ttl=600)
def carica_db_ottimizzato():
    try:
        # Legge usando il TTL interno della connessione
        df = conn.read(ttl="10m")
        
        if df is None or df.empty:
            return pd.DataFrame(columns=["Nominativo", "Tipo", "Ruolo", "Maglia", "GG", "MM", "AA", "FIGC", "Capitano", "Portiere"])

        # Pulizia nomi colonne
        df.columns = [str(c).strip() for c in df.columns]
        
        # Garanzia colonne presenti
        col_richieste = ["Nominativo", "Tipo", "Ruolo", "Maglia", "GG", "MM", "AA", "FIGC", "Capitano", "Portiere"]
        for col in col_richieste:
            if col not in df.columns:
                df[col] = False if col in ["Capitano", "Portiere"] else ""
        
        # Formattazione dati
        df['Tipo'] = df['Tipo'].astype(str).str.strip().fillna("Giocatore")
        df['Nominativo'] = df['Nominativo'].astype(str).str.strip()
        df['Maglia'] = pd.to_numeric(df['Maglia'], errors='coerce')
        
        for c in ["Capitano", "Portiere"]:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0).astype(bool)
            
        return df.sort_values(by=['Tipo', 'Nominativo'], ascending=[False, True])
    except Exception as e:
        st.error(f"Errore nel caricamento: {e}")
        return pd.DataFrame()

def salva_db(df):
    try:
        conn.update(data=df)
        # SVUOTA LA CACHE DOPO IL SALVATAGGIO
        # Così i dati nuovi saranno subito visibili
        st.cache_data.clear()
        return True
    except Exception as e:
        st.error(f"Errore nel salvataggio: {e}")
        return False

# --- LOGICA EXCEL (Invariata) ---
def safe_write(ws, cell_coord, value):
    from openpyxl.cell.cell import MergedCell
    cell = ws[cell_coord]
    if isinstance(cell, MergedCell):
        for m_range in ws.merged_cells.ranges:
            if cell_coord in m_range:
                ws.cell(row=m_range.min_row, column=m_range.min_col).value = value
                return
    else:
        cell.value = value

def compila_template(p_df, s_df, info):
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active 
    safe_write(ws, 'G7', f"Zenith Prato Vs {info['avversario']}")
    safe_write(ws, 'G8', f"Data: {info['data']} - Ora: {info['ora']}")
    safe_write(ws, 'G9', info['campo'])
    p_df = p_df.sort_values(by='Maglia', ascending=True)

    for i, (_, row) in enumerate(p_df.iterrows()):
        r = 12 + i
        if r > 38: break
        nome = f"{row.get('Nominativo', '')}"
        if row.get('Capitano'): nome += " (C)"
        if row.get('Portiere'): nome += " (P)"
        safe_write(ws, f'C{r}', row.get('Maglia', ''))
        safe_write(ws, f'D{r}', row.get('GG', ''))
        safe_write(ws, f'E{r}', row.get('MM', ''))
        safe_write(ws, f'F{r}', row.get('AA', ''))
        safe_write(ws, f'G{r}', nome)
        safe_write(ws, f'I{r}', row.get('FIGC', ''))

    for i, (_, row) in enumerate(s_df.iterrows()):
        r = 39 + i
        safe_write(ws, f'C{r}', row.get('Ruolo', ''))       
        safe_write(ws, f'D{r}', row.get('Nominativo', ''))  
        safe_write(ws, f'I{r}', row.get('FIGC', ''))        

    out = BytesIO()
    wb.save(out)
    return out.getvalue()

# --- INTERFACCIA ---
st.title("⚽ Zenith Prato - Sistema Distinte")
t1, t2 = st.tabs(["📋 Genera Distinta", "⚙️ Gestione Anagrafica"])

with t2:
    st.header("Anagrafica (Dati in Cache)")
    # Usa la funzione ottimizzata
    df = carica_db_ottimizzato()
    
    if not df.empty:
        giocatori_df = df[df['Tipo'].str.contains('giocatore', case=False, na=False)]
        giocatori_nomi = giocatori_df['Nominativo'].tolist()
        
        st.subheader("🏆 Ruoli Speciali")
        c1, c2 = st.columns(2)
        with c1:
            cap_lista = df[df['Capitano'] == True]['Nominativo'].tolist()
            idx_cap = (giocatori_nomi.index(cap_lista[0]) + 1) if cap_lista and cap_lista[0] in giocatori_nomi else 0
            capitano = st.selectbox("Seleziona Capitano", ["Nessuno"] + giocatori_nomi, index=idx_cap)
        with c2:
            por_lista = df[df['Portiere'] == True]['Nominativo'].tolist()
            idx_por = (giocatori_nomi.index(por_lista[0]) + 1) if por_lista and por_lista[0] in giocatori_nomi else 0
            portiere = st.selectbox("Seleziona Portiere", ["Nessuno"] + giocatori_nomi, index=idx_por)

        col_vis = ["Nominativo", "Tipo", "Ruolo", "Maglia", "GG", "MM", "AA", "FIGC"]
        df_edit = st.data_editor(df[col_vis], num_rows="dynamic", use_container_width=True, hide_index=True)
        
        if st.button("💾 Salva e Aggiorna Cache"):
            df_edit['Capitano'] = df_edit['Nominativo'] == capitano
            df_edit['Portiere'] = df_edit['Nominativo'] == portiere
            if salva_db(df_edit):
                st.success("Dati salvati e cache svuotata!")
                st.rerun()

with t1:
    st.header("Gara")
    c1, c2 = st.columns(2)
    with c1:
        avv = st.text_input("Avversario", "SQUADRA OSPITE")
        cmp = st.text_input("Campo", "Chiavacci")
    with c2:
        dat = st.text_input("Data", "15/04/2026")
        ora = st.text_input("Ora", "10:30")

    df_gara = carica_db_ottimizzato()
    if not df_gara.empty:
        gioc = df_gara[df_gara['Tipo'].str.contains('giocatore', case=False, na=False)]
        staf = df_gara[df_gara['Tipo'].str.contains('staff', case=False, na=False)]
        
        col_s1, col_s2 = st.columns(2)
        with col_s1:
            sel_p = st.multiselect("Giocatori", gioc['Nominativo'].tolist())
        with col_s2:
            sel_s = st.multiselect("Staff", staf['Nominativo'].tolist())

        if st.button("🚀 Genera Excel"):
            if sel_p:
                xlsx = compila_template(
                    gioc[gioc['Nominativo'].isin(sel_p)], 
                    staf[staf['Nominativo'].isin(sel_s)], 
                    {"avversario": avv, "campo": cmp, "data": dat, "ora": ora}
                )
                st.download_button("📥 Scarica", xlsx, f"Distinta_{avv}.xlsx")
